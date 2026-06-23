"""Load farmer-facing fixed strings from translated_languages.xlsx by (script, vocal)."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

OFFICIAL_LANGUAGES = [
    "Assamese",
    "Bengali",
    "Bodo",
    "Dogri",
    "Gujarati",
    "Hindi",
    "Kannada",
    "Kashmiri",
    "Konkani",
    "Maithili",
    "Malayalam",
    "Manipuri (Meitei)",
    "Marathi",
    "Nepali",
    "Odia",
    "Punjabi",
    "Sanskrit",
    "Santali",
    "Sindhi",
    "Tamil",
    "Telugu",
    "Urdu",
    "English",
]

_DEFAULT_SCRIPT = "English"
_DEFAULT_VOCAL = "English"

_CATALOG_PATH = Path(__file__).resolve().parent / "translated_languages.xlsx"

_COL_SCRIPT = "Script Language"
_COL_VOCAL = "Vocal Language"
_COL_TWO_HOUR = "2 hour disclaimer"
_COL_STATE = "State Follow Up"
_COL_CROP = "Crop Follow Up"
_COL_TESTING = "Testing disclaimer"
_COL_LATE_NIGHT = "Questions submitted between 10:01 PM and 11:59 PM"
_COL_EARLY_MORNING = "Questions submitted between 12:00 AM and 5:59 AM"


@dataclass(frozen=True)
class CatalogRow:
    script_language: str
    vocal_language: str
    two_hour_disclaimer: str
    state_follow_up: str
    crop_follow_up: str
    testing_disclaimer: str
    late_night_disclaimer: str
    early_morning_disclaimer: str


def _normalize_lang(name: str) -> str:
    return (name or "").strip()


def _lang_key(script: str, vocal: str) -> tuple[str, str]:
    return (_normalize_lang(script), _normalize_lang(vocal))


def load_catalog(path: Optional[Path] = None) -> dict[tuple[str, str], CatalogRow]:
    """Load all rows from the Excel catalog."""
    from openpyxl import load_workbook

    xlsx_path = path or _CATALOG_PATH
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]

    def col_index(name: str) -> int:
        for i, h in enumerate(header):
            if h == name:
                return i
        raise ValueError(f"Column {name!r} not found in {xlsx_path}")

    idx_script = col_index(_COL_SCRIPT)
    idx_vocal = col_index(_COL_VOCAL)
    idx_two = col_index(_COL_TWO_HOUR)
    idx_state = col_index(_COL_STATE)
    idx_crop = col_index(_COL_CROP)
    idx_testing = col_index(_COL_TESTING)
    idx_late_night = col_index(_COL_LATE_NIGHT)
    idx_early_morning = col_index(_COL_EARLY_MORNING)

    catalog: dict[tuple[str, str], CatalogRow] = {}
    for row in rows_iter:
        if not row or row[idx_script] is None or row[idx_vocal] is None:
            continue
        script = _normalize_lang(str(row[idx_script]))
        vocal = _normalize_lang(str(row[idx_vocal]))
        if not script or not vocal:
            continue

        def cell(i: int) -> str:
            v = row[i] if i < len(row) else None
            return str(v).strip() if v is not None else ""

        catalog[_lang_key(script, vocal)] = CatalogRow(
            script_language=script,
            vocal_language=vocal,
            two_hour_disclaimer=cell(idx_two),
            state_follow_up=cell(idx_state),
            crop_follow_up=cell(idx_crop),
            testing_disclaimer=cell(idx_testing),
            late_night_disclaimer=cell(idx_late_night),
            early_morning_disclaimer=cell(idx_early_morning),
        )
    wb.close()
    return catalog


_catalog_cache: Optional[dict[tuple[str, str], CatalogRow]] = None


def get_catalog() -> dict[tuple[str, str], CatalogRow]:
    global _catalog_cache
    if _catalog_cache is None:
        _catalog_cache = load_catalog()
    return _catalog_cache


def get_catalog_row(
    script_language: str,
    vocal_language: str,
    *,
    catalog: Optional[dict[tuple[str, str], CatalogRow]] = None,
) -> CatalogRow:
    """Lookup by (script, vocal); fallback to English/English."""
    cat = catalog if catalog is not None else get_catalog()
    script = _normalize_lang(script_language) or _DEFAULT_SCRIPT
    vocal = _normalize_lang(vocal_language) or _DEFAULT_VOCAL
    row = cat.get(_lang_key(script, vocal))
    if row is not None:
        return row
    fallback = cat.get(_lang_key(_DEFAULT_SCRIPT, _DEFAULT_VOCAL))
    if fallback is not None:
        logger.warning(
            "translation_catalog: missing (%s, %s) — using English/English",
            script,
            vocal,
        )
        return fallback
    raise KeyError(f"No catalog row for ({script}, {vocal}) and no English/English fallback")


def get_testing_disclaimer(script_language: str, vocal_language: str) -> str:
    return get_catalog_row(script_language, vocal_language).testing_disclaimer


def get_two_hour_disclaimer(script_language: str, vocal_language: str) -> str:
    return get_catalog_row(script_language, vocal_language).two_hour_disclaimer


def get_state_follow_up(script_language: str, vocal_language: str) -> str:
    return get_catalog_row(script_language, vocal_language).state_follow_up


def get_crop_follow_up(script_language: str, vocal_language: str) -> str:
    return get_catalog_row(script_language, vocal_language).crop_follow_up


def get_late_night_disclaimer(script_language: str, vocal_language: str) -> str:
    """Get the late night disclaimer (10:01 PM - 11:59 PM) for the given language pair."""
    return get_catalog_row(script_language, vocal_language).late_night_disclaimer


def get_early_morning_disclaimer(script_language: str, vocal_language: str) -> str:
    """Get the early morning disclaimer (12:00 AM - 5:59 AM) for the given language pair."""
    return get_catalog_row(script_language, vocal_language).early_morning_disclaimer


def language_pair_from_plan(plan: Optional[dict]) -> tuple[str, str]:
    """Default script/vocal from planner plan."""
    p = plan or {}
    script = _normalize_lang(p.get("script_language") or "") or _DEFAULT_SCRIPT
    vocal = _normalize_lang(p.get("vocal_language") or "") or _DEFAULT_VOCAL
    return script, vocal


def synthesis_lang_label(script_language: str, vocal_language: str) -> str:
    """Legacy label for GDB source prefix helpers (script + vocal)."""
    script = _normalize_lang(script_language) or _DEFAULT_SCRIPT
    vocal = _normalize_lang(vocal_language) or _DEFAULT_VOCAL
    if script == "English" and vocal == "English":
        return "English"
    if script == "English" and vocal == "Hindi":
        return "Hinglish"
    if script == vocal:
        return vocal
    return f"Romanized {vocal}"


def needs_translation(script_language: str, vocal_language: str) -> bool:
    """True when a post-synthesis translate step is required."""
    script, vocal = (
        _normalize_lang(script_language) or _DEFAULT_SCRIPT,
        _normalize_lang(vocal_language) or _DEFAULT_VOCAL,
    )
    return not (script == "English" and vocal == "English")
